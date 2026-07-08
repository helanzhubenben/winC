import csv
import json
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from django.http import HttpResponse
from django.db import IntegrityError
from django.db.models import Q, Sum
from django.db.models.functions import Coalesce, ExtractYear
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action, api_view
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from .leveling import recalculate_customer_levels
from .locations import normalize_area_text
from .models import (
    Customer, Contact, CustomerContactRecord, CustomerRevenue, WeeklyReport, get_last_quarter_range
)
from .serializers import (
    CustomerSerializer, CustomerListSerializer, ContactSerializer,
    CustomerRevenueSerializer, WeeklyReportSerializer, WeeklyReportListSerializer
)


REVENUE_REQUIRED_COLUMNS = {'month', 'customer name', 'revenue'}
EXCEL_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
CUSTOMER_IMPORT_SHEET_NAME = '客户导入'
CUSTOMER_IMPORT_COLUMNS = [
    '客户名称',
    '别名',
    '业务模式',
    '区域',
    '城市',
    '地址',
    'X轴描述',
    'X轴评分',
    'Y轴描述',
    'Y轴评分',
    'Z轴描述',
    'Z轴评分',
    '客户策略',
    '潜在贡献',
    '备注',
    '联系人姓名',
    '联系人职位',
    '联系人电话',
    '联系人邮箱',
    '是否关键人',
]
CUSTOMER_IMPORT_REQUIRED_COLUMNS = [column for column in CUSTOMER_IMPORT_COLUMNS if column != '别名']
CUSTOMER_IMPORT_BUSINESS_MODELS = {'Hunting', 'Farming'}
CUSTOMER_IMPORT_TRUE_VALUES = {'是', 'y', 'yes', 'true', '1'}
CUSTOMER_IMPORT_FALSE_VALUES = {'否', 'n', 'no', 'false', '0'}
CUSTOMER_FILTER_FIELDS = {
    'client_name': {'type': 'text', 'field': 'client_name'},
    'alias': {'type': 'text', 'field': 'alias'},
    'business_model': {'type': 'choice', 'field': 'business_model'},
    'area': {'type': 'text', 'field': 'area'},
    'city': {'type': 'text', 'field': 'city'},
    'level': {'type': 'choice', 'field': 'level'},
    'status': {'type': 'choice', 'field': 'status'},
    'address': {'type': 'text', 'field': 'address'},
    'remark': {'type': 'text', 'field': 'remark'},
    'score_x': {'type': 'number', 'field': 'score_x'},
    'score_y': {'type': 'number', 'field': 'score_y'},
    'score_z': {'type': 'number', 'field': 'score_z'},
    'potential_contribution': {'type': 'number', 'field': 'potential_contribution'},
    'created_at': {'type': 'date', 'field': 'created_at'},
    'updated_at': {'type': 'date', 'field': 'updated_at'},
    'last_year_revenue': {'type': 'computed_revenue', 'annotation': 'filtered_last_year_revenue'},
    'last_quarter_revenue': {'type': 'computed_revenue', 'annotation': 'filtered_last_quarter_revenue'},
}


def _xlsx_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)

    response = HttpResponse(output.getvalue(), content_type=EXCEL_CONTENT_TYPE)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _repo_root():
    return Path(__file__).resolve().parents[2]


def _customer_import_template_path():
    return _repo_root() / 'docs' / 'templates' / 'customer_import_template.xlsx'


def _normalize_header(value):
    return str(value or '').strip().lower()


def _blank(value):
    return value is None or str(value).strip() == ''


def _trimmed(value):
    return str(value or '').strip()


def _format_export_date(value):
    return value.strftime('%Y-%m-%d') if value else ''


def _format_export_datetime(value):
    return timezone.localtime(value).strftime('%Y-%m-%d %H:%M:%S') if value else ''


def _last_year_range():
    today = timezone.localdate()
    start_date = today.replace(year=today.year - 1, month=1, day=1)
    end_date = today.replace(year=today.year - 1, month=12, day=31)
    return start_date, end_date


def _parse_filter_decimal(value):
    if _blank(value):
        return None
    try:
        return Decimal(str(value).strip().replace(',', ''))
    except InvalidOperation:
        return None


def _add_customers_sheet(workbook, queryset=None):
    sheet = workbook.active
    sheet.title = '客户池'
    sheet.append([
        '客户名称',
        '别名',
        '业务模式',
        '区域',
        '城市',
        '客户等级',
        'X评分',
        'Y评分',
        'Z评分',
        '潜在贡献',
        '去年营收',
        '上季度营收',
        '联系人数',
        '更新时间',
    ])

    queryset = queryset or Customer.objects.all().prefetch_related('contacts', 'revenue_records')
    serializer = CustomerListSerializer(queryset, many=True)
    for item in serializer.data:
        sheet.append([
            item.get('name', ''),
            item.get('alias', ''),
            item.get('business_model', ''),
            item.get('region', ''),
            item.get('city', ''),
            item.get('level', ''),
            item.get('score_x', 0),
            item.get('score_y', 0),
            item.get('score_z', 0),
            item.get('potential_contribution') or '',
            item.get('last_year_revenue', '0.00'),
            item.get('last_quarter_revenue', '0.00'),
            item.get('contacts_count', 0),
            item.get('updated_at', ''),
        ])
    return sheet


def _action_value(action, key, fallback=''):
    if isinstance(action, dict):
        return action.get(key, fallback)
    return fallback


def _add_weekly_reports_sheet(workbook, queryset=None):
    sheet = workbook.active if workbook.sheetnames == ['Sheet'] else workbook.create_sheet('周报')
    sheet.title = '周报'
    sheet.append([
        '客户名称',
        '匹配客户',
        '区域',
        '地址',
        '任务',
        '项目定义',
        '状态',
        '到期日期',
        '修订日期',
        '完成日期',
        '营收',
        '责任人',
        '备注',
        'Action序号',
        'Action日期',
        'Action内容',
        'Action结果',
        'Action下一步',
        'Action用户',
        'Action创建时间',
        '更新时间',
    ])

    queryset = queryset or WeeklyReport.objects.all().select_related('customer')
    for report in queryset:
        actions = report.actions or [None]
        for index, action in enumerate(actions, start=1):
            action_content = _action_value(action, 'action') or _action_value(action, 'content')
            sheet.append([
                report.client_name,
                report.customer.client_name if report.customer else '',
                report.area,
                report.address,
                report.tasks,
                report.definition,
                '已完成' if report.status == 'completed' else '进行中',
                _format_export_date(report.due_date),
                _format_export_date(report.revise_date),
                _format_export_date(report.finish_date),
                report.revenue,
                report.responsibility,
                report.remark,
                index if action else '',
                _action_value(action, 'action_date'),
                action_content,
                _action_value(action, 'result'),
                _action_value(action, 'next_step'),
                _action_value(action, 'user'),
                _action_value(action, 'timestamp'),
                _format_export_datetime(report.updated_at),
            ])
    return sheet


@api_view(['GET'])
def export_workbook(request):
    workbook = Workbook()
    _add_customers_sheet(workbook)
    _add_weekly_reports_sheet(workbook)
    return _xlsx_response(workbook, 'fishpool-export.xlsx')


def _read_revenue_rows(upload):
    filename = upload.name.lower()
    if filename.endswith('.xlsx'):
        return _read_xlsx_revenue_rows(upload)
    if filename.endswith('.csv'):
        return _read_csv_revenue_rows(upload)
    raise ValueError('仅支持 .xlsx 或 .csv 文件')


def _read_csv_revenue_rows(upload):
    content = upload.read().decode('utf-8-sig')
    reader = csv.DictReader(StringIO(content))
    if not reader.fieldnames:
        raise ValueError('文件缺少表头')

    headers = {_normalize_header(name): name for name in reader.fieldnames}
    missing = REVENUE_REQUIRED_COLUMNS - set(headers.keys())
    if missing:
        raise ValueError(f"缺少字段: {', '.join(sorted(missing))}")

    rows = []
    for index, row in enumerate(reader, start=2):
        rows.append((
            index,
            {column: row.get(headers[column]) for column in REVENUE_REQUIRED_COLUMNS}
        ))
    return rows


def _read_xlsx_revenue_rows_active_sheet(upload):
    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError('文件缺少表头')

    headers = {_normalize_header(name): position for position, name in enumerate(header_row)}
    missing = REVENUE_REQUIRED_COLUMNS - set(headers.keys())
    if missing:
        raise ValueError(f"缺少字段: {', '.join(sorted(missing))}")

    rows = []
    for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        rows.append((
            index,
            {column: values[headers[column]] if headers[column] < len(values) else None for column in REVENUE_REQUIRED_COLUMNS}
        ))
    workbook.close()
    return rows


def _read_xlsx_revenue_rows(upload):
    workbook = load_workbook(upload, read_only=True, data_only=True)
    rows = []
    missing_error = None
    found_valid_sheet = False

    try:
        for sheet in workbook.worksheets:
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue

            headers = {_normalize_header(name): position for position, name in enumerate(header_row)}
            missing = REVENUE_REQUIRED_COLUMNS - set(headers.keys())
            if missing:
                missing_error = f"缺少字段: {', '.join(sorted(missing))}"
                continue

            found_valid_sheet = True
            for index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                rows.append((
                    index,
                    {
                        column: values[headers[column]] if headers[column] < len(values) else None
                        for column in REVENUE_REQUIRED_COLUMNS
                    }
                ))
    finally:
        workbook.close()

    if found_valid_sheet:
        return rows
    if missing_error:
        raise ValueError(missing_error)
    raise ValueError('文件缺少表头')


def _parse_month(value):
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)

    raw_value = str(value or '').strip()
    if not raw_value:
        raise ValueError('month 不能为空')

    for date_format in ('%Y-%m', '%Y/%m', '%Y.%m', '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d'):
        try:
            parsed = datetime.strptime(raw_value, date_format).date()
            return parsed.replace(day=1)
        except ValueError:
            continue
    raise ValueError('month 格式不正确，请使用 YYYY-MM 或 YYYY-MM-DD')


def _parse_revenue(value):
    raw_value = str(value or '').strip().replace(',', '')
    if not raw_value:
        raise ValueError('revenue 不能为空')
    try:
        return Decimal(raw_value)
    except InvalidOperation as exc:
        raise ValueError('revenue 必须是数字') from exc


def _read_customer_import_rows(upload):
    if not upload.name.lower().endswith('.xlsx'):
        raise ValueError('仅支持 .xlsx 文件')

    try:
        workbook = load_workbook(upload, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('文件无法读取，请确认是有效的 .xlsx 文件') from exc

    sheet = workbook[CUSTOMER_IMPORT_SHEET_NAME] if CUSTOMER_IMPORT_SHEET_NAME in workbook.sheetnames else workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        workbook.close()
        raise ValueError('文件缺少表头')

    headers = {_trimmed(name): index for index, name in enumerate(header_row) if not _blank(name)}
    missing = [column for column in CUSTOMER_IMPORT_REQUIRED_COLUMNS if column not in headers]
    if missing:
        workbook.close()
        raise ValueError(f"缺少字段: {', '.join(missing)}")

    rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(_blank(value) for value in values):
            continue
        rows.append((
            row_number,
            {
                column: (
                    values[headers[column]]
                    if column in headers and headers[column] < len(values)
                    else None
                )
                for column in CUSTOMER_IMPORT_COLUMNS
            }
        ))

    workbook.close()
    return rows


def _parse_required_text(row, column):
    value = _trimmed(row.get(column))
    if not value:
        raise ValueError(f'{column}不能为空')
    return value


def _parse_optional_text(row, column):
    return _trimmed(row.get(column))


def _parse_score(row, column):
    raw_value = row.get(column)
    if _blank(raw_value):
        return 0
    try:
        score = Decimal(str(raw_value).strip())
    except InvalidOperation as exc:
        raise ValueError(f'{column}必须是 0-100 的数字') from exc
    if score < 0 or score > 100:
        raise ValueError(f'{column}必须是 0-100 的数字')
    if score != score.to_integral_value():
        raise ValueError(f'{column}必须是整数')
    return int(score)


def _parse_optional_decimal(row, column):
    raw_value = row.get(column)
    if _blank(raw_value):
        return None
    try:
        return Decimal(str(raw_value).strip().replace(',', ''))
    except InvalidOperation as exc:
        raise ValueError(f'{column}必须是数字') from exc


def _parse_boolean(row, column):
    raw_value = row.get(column)
    if _blank(raw_value):
        return False

    value = str(raw_value).strip().lower()
    if value in CUSTOMER_IMPORT_TRUE_VALUES:
        return True
    if value in CUSTOMER_IMPORT_FALSE_VALUES:
        return False
    raise ValueError(f'{column}必须是 是/否、Y/N、true/false 或 1/0')


def _parse_customer_import_row(row):
    business_model = _parse_required_text(row, '业务模式')
    if business_model not in CUSTOMER_IMPORT_BUSINESS_MODELS:
        raise ValueError('业务模式必须是 Hunting 或 Farming')

    score_x = _parse_score(row, 'X轴评分')
    score_y = _parse_score(row, 'Y轴评分')
    score_z = _parse_score(row, 'Z轴评分')

    customer_name = _parse_required_text(row, '客户名称')
    customer_data = {
        'alias': _parse_optional_text(row, '别名'),
        'business_model': business_model,
        'area': normalize_area_text(row.get('区域')),
        'city': _parse_optional_text(row, '城市'),
        'address': _parse_optional_text(row, '地址'),
        'description_x': _parse_optional_text(row, 'X轴描述'),
        'score_x': score_x,
        'description_y': _parse_optional_text(row, 'Y轴描述'),
        'score_y': score_y,
        'key_person': _parse_optional_text(row, 'Z轴描述'),
        'score_z': score_z,
        'client_strategy': _parse_optional_text(row, '客户策略'),
        'potential_contribution': _parse_optional_decimal(row, '潜在贡献'),
        'remark': _parse_optional_text(row, '备注'),
    }

    contact_name = _parse_optional_text(row, '联系人姓名')
    contact_data = None
    if contact_name:
        contact_data = {
            'name': contact_name,
            'position': _parse_optional_text(row, '联系人职位'),
            'phone': _parse_optional_text(row, '联系人电话'),
            'email': _parse_optional_text(row, '联系人邮箱'),
            'is_key_person': _parse_boolean(row, '是否关键人'),
        }

    return customer_name, customer_data, contact_data


def _customer_data_changed(first_data, current_data):
    return any(first_data.get(key) != current_data.get(key) for key in first_data.keys())


def _import_customer_and_contact_rows(rows):
    customers_created = 0
    customers_updated = 0
    contacts_created = 0
    contacts_updated = 0
    skipped = 0
    errors = []
    warnings = []
    customer_cache = {}
    first_customer_data = {}
    warned_customers = set()

    for row_number, row in rows:
        try:
            customer_name, customer_data, contact_data = _parse_customer_import_row(row)
        except ValueError as exc:
            skipped += 1
            errors.append({
                'row': row_number,
                'customer_name': _trimmed(row.get('客户名称')),
                'contact_name': _trimmed(row.get('联系人姓名')),
                'reason': str(exc),
            })
            continue

        if customer_name in customer_cache:
            customer = customer_cache[customer_name]
            if (
                customer_name not in warned_customers and
                _customer_data_changed(first_customer_data[customer_name], customer_data)
            ):
                warnings.append({
                    'row': row_number,
                    'customer_name': customer_name,
                    'reason': '重复客户行的客户字段与第一次出现不一致，已保留第一次出现的客户字段',
                })
                warned_customers.add(customer_name)
        else:
            customer = Customer.objects.filter(client_name__iexact=customer_name).first()
            created = customer is None
            if created:
                customer = Customer.objects.create(client_name=customer_name, **customer_data)
            else:
                for field, value in customer_data.items():
                    setattr(customer, field, value)
                customer.save()
            customer_cache[customer_name] = customer
            first_customer_data[customer_name] = customer_data
            if created:
                customers_created += 1
            else:
                customers_updated += 1

        if contact_data:
            contact = Contact.objects.filter(customer=customer, name=contact_data['name']).first()
            created = contact is None
            if created:
                Contact.objects.create(customer=customer, **contact_data)
            else:
                contact.position = contact_data['position']
                contact.phone = contact_data['phone']
                contact.email = contact_data['email']
                contact.is_key_person = contact_data['is_key_person']
                contact.save()
            if created:
                contacts_created += 1
            else:
                contacts_updated += 1

    recalculate_customer_levels()

    return {
        'customers_created': customers_created,
        'customers_updated': customers_updated,
        'contacts_created': contacts_created,
        'contacts_updated': contacts_updated,
        'skipped': skipped,
        'errors': errors[:50],
        'warnings': warnings[:50],
    }


class CustomerViewSet(viewsets.ModelViewSet):
    """客户视图集"""

    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['level', 'area', 'city', 'business_model', 'status']
    search_fields = ['client_name', 'alias', 'address', 'remark']
    ordering_fields = [
        'client_name',
        'alias',
        'business_model',
        'area',
        'city',
        'level',
        'status',
        'score_x',
        'score_y',
        'score_z',
        'potential_contribution',
        'created_at',
        'updated_at',
        'last_year_revenue',
        'last_quarter_revenue',
    ]
    ordering = ['-updated_at']

    def get_serializer_class(self):
        """根据action选择序列化器"""
        if self.action == 'list':
            return CustomerListSerializer
        return CustomerSerializer

    def perform_create(self, serializer):
        customer = serializer.save()
        recalculate_customer_levels()
        customer.refresh_from_db()

    def perform_update(self, serializer):
        customer = serializer.save()
        recalculate_customer_levels()
        customer.refresh_from_db()

    def get_queryset(self):
        """优化查询性能"""
        queryset = super().get_queryset()
        queryset = self._apply_custom_filters(queryset)
        queryset = self._annotate_custom_ordering(queryset)
        if self.action == 'list':
            queryset = queryset.prefetch_related('contacts', 'revenue_records', 'contact_records')
        elif self.action == 'retrieve':
            queryset = queryset.prefetch_related('contacts', 'revenue_records', 'contact_records')
        return queryset

    def _annotate_custom_ordering(self, queryset):
        ordering = _trimmed(self.request.query_params.get('ordering')).lstrip('-')
        if ordering == 'last_year_revenue':
            start_date, end_date = _last_year_range()
            return queryset.annotate(last_year_revenue=self._revenue_sum_annotation(start_date, end_date))
        if ordering == 'last_quarter_revenue':
            start_date, end_date = get_last_quarter_range()
            return queryset.annotate(last_quarter_revenue=self._revenue_sum_annotation(start_date, end_date))
        return queryset

    def _apply_custom_filters(self, queryset):
        raw_filters = self.request.query_params.get('filters')
        if not raw_filters:
            return queryset

        try:
            filters_data = json.loads(raw_filters)
        except (TypeError, ValueError):
            return queryset
        if not isinstance(filters_data, list):
            return queryset

        needs_last_year = any(item.get('field') == 'last_year_revenue' for item in filters_data if isinstance(item, dict))
        needs_last_quarter = any(item.get('field') == 'last_quarter_revenue' for item in filters_data if isinstance(item, dict))
        if needs_last_year or needs_last_quarter:
            queryset = self._annotate_revenue_filters(queryset, needs_last_year, needs_last_quarter)

        for item in filters_data:
            if not isinstance(item, dict):
                continue
            field_name = item.get('field')
            operator = item.get('operator')
            value = item.get('value')
            field_config = CUSTOMER_FILTER_FIELDS.get(field_name)
            if not field_config:
                continue
            queryset = self._apply_custom_filter(queryset, field_config, operator, value)
        return queryset

    def _annotate_revenue_filters(self, queryset, needs_last_year, needs_last_quarter):
        annotations = {}
        if needs_last_year:
            start_date, end_date = _last_year_range()
            annotations['filtered_last_year_revenue'] = self._revenue_sum_annotation(start_date, end_date)
        if needs_last_quarter:
            start_date, end_date = get_last_quarter_range()
            annotations['filtered_last_quarter_revenue'] = self._revenue_sum_annotation(start_date, end_date)
        return queryset.annotate(**annotations) if annotations else queryset

    def _revenue_sum_annotation(self, start_date, end_date):
        return Coalesce(
            Sum(
                'revenue_records__revenue',
                filter=Q(revenue_records__month__gte=start_date, revenue_records__month__lte=end_date),
                default=Decimal('0'),
            ),
            Decimal('0'),
        )

    def _apply_custom_filter(self, queryset, field_config, operator, value):
        field_type = field_config['type']
        field = field_config.get('field') or field_config.get('annotation')

        if field_type == 'text':
            text = _trimmed(value)
            if not text:
                return queryset
            if operator == 'eq':
                return queryset.filter(**{f'{field}__iexact': text})
            return queryset.filter(**{f'{field}__icontains': text})

        if field_type == 'choice':
            text = _trimmed(value)
            if not text:
                return queryset
            return queryset.filter(**{field: text})

        if field_type in {'number', 'computed_revenue'}:
            return self._apply_decimal_filter(queryset, field, operator, value)

        if field_type == 'date':
            return self._apply_date_filter(queryset, field, operator, value)

        return queryset

    def _apply_decimal_filter(self, queryset, field, operator, value):
        if operator == 'between' and isinstance(value, list) and len(value) >= 2:
            low = _parse_filter_decimal(value[0])
            high = _parse_filter_decimal(value[1])
            if low is not None:
                queryset = queryset.filter(**{f'{field}__gte': low})
            if high is not None:
                queryset = queryset.filter(**{f'{field}__lte': high})
            return queryset

        number = _parse_filter_decimal(value)
        if number is None:
            return queryset
        if operator == 'lte':
            return queryset.filter(**{f'{field}__lte': number})
        if operator == 'eq':
            return queryset.filter(**{field: number})
        return queryset.filter(**{f'{field}__gte': number})

    def _apply_date_filter(self, queryset, field, operator, value):
        if operator == 'between' and isinstance(value, list) and len(value) >= 2:
            start_date = parse_date(_trimmed(value[0]))
            end_date = parse_date(_trimmed(value[1]))
            if start_date:
                queryset = queryset.filter(**{f'{field}__date__gte': start_date})
            if end_date:
                queryset = queryset.filter(**{f'{field}__date__lte': end_date})
            return queryset

        parsed = parse_date(_trimmed(value))
        if not parsed:
            return queryset
        if operator == 'lte':
            return queryset.filter(**{f'{field}__date__lte': parsed})
        if operator == 'eq':
            return queryset.filter(**{f'{field}__date': parsed})
        return queryset.filter(**{f'{field}__date__gte': parsed})

    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """获取统计数据"""
        queryset = self.filter_queryset(self.get_queryset())

        stats = {
            'total': queryset.count(),
            'by_level': {},
            'by_business_model': {},
            'by_area': {},
        }

        # 按等级统计
        for level in ['A', 'B', 'C', 'D', 'X']:
            stats['by_level'][level] = queryset.filter(level=level).count()

        # 按业务模式统计
        for model in ['Hunting', 'Farming']:
            stats['by_business_model'][model] = queryset.filter(business_model=model).count()

        # 按区域统计（前10）
        areas = queryset.values('area').distinct()
        for area_dict in areas[:10]:
            area = area_dict['area']
            if area:
                stats['by_area'][area] = queryset.filter(area=area).count()

        return Response(stats)

    @action(detail=True, methods=['get'], url_path='weekly-reports')
    def weekly_reports(self, request, pk=None):
        """获取客户的所有 Weekly Report"""
        customer = self.get_object()
        reports = customer.weekly_reports.all().order_by('-due_date', '-created_at')
        serializer = WeeklyReportListSerializer(reports, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'], url_path='revenue-summary')
    def revenue_summary(self, request, pk=None):
        """Return yearly revenue totals and monthly revenue rows for the customer."""
        customer = self.get_object()
        records = CustomerRevenue.objects.filter(customer=customer)

        yearly = [
            {
                'year': item['year'],
                'revenue': f'{item["total"] or Decimal("0"):.2f}',
            }
            for item in records
            .annotate(year=ExtractYear('month'))
            .values('year')
            .annotate(total=Sum('revenue'))
            .order_by('year')
        ]

        monthly_records = records.order_by('month')
        year = _trimmed(request.query_params.get('year'))
        if year.isdigit():
            monthly_records = monthly_records.filter(month__year=int(year))

        monthly = [
            {
                'month': record.month.strftime('%Y-%m'),
                'revenue': f'{record.revenue:.2f}',
            }
            for record in monthly_records
        ]

        return Response({
            'yearly': yearly,
            'monthly': monthly,
        })

    @action(detail=True, methods=['post'], url_path='record-contact')
    def record_contact(self, request, pk=None):
        """Record one customer contact timestamp; each customer can be recorded once per local day."""
        customer = self.get_object()
        now = timezone.now()
        today = timezone.localdate(now)
        latest = customer.contact_records.order_by('-contacted_at').first()

        if customer.contact_records.filter(contact_date=today).exists():
            return Response({
                'error': '今天已经记录过联系，明天 00:00 后可再次记录',
                'last_contacted_at': latest.contacted_at.isoformat() if latest else None,
                'contacted_today': True,
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            record = CustomerContactRecord.objects.create(
                customer=customer,
                contacted_at=now,
            )
        except IntegrityError:
            return Response({
                'error': '今天已经记录过联系，明天 00:00 后可再次记录',
                'last_contacted_at': latest.contacted_at.isoformat() if latest else None,
                'contacted_today': True,
            }, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'id': record.id,
            'last_contacted_at': record.contacted_at.isoformat(),
            'contacted_today': record.contact_date == today,
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'], url_path='create-action')
    def create_action(self, request, pk=None):
        """Create a Weekly Report action from a customer context."""
        customer = self.get_object()
        definition = _trimmed(request.data.get('definition'))
        action_content = _trimmed(request.data.get('action') or request.data.get('content'))

        if not definition:
            return Response({'error': '项目定义不能为空'}, status=status.HTTP_400_BAD_REQUEST)
        if not action_content:
            return Response({'error': '行动内容不能为空'}, status=status.HTTP_400_BAD_REQUEST)

        action_record = {
            'timestamp': timezone.now().isoformat(),
            'action_date': _trimmed(request.data.get('action_date')),
            'action': action_content,
            'result': _trimmed(request.data.get('result')),
            'next_step': _trimmed(request.data.get('next_step')),
            'user': _trimmed(request.data.get('user')),
        }
        report_data = {
            'client_name': customer.client_name,
            'customer': customer.id,
            'area': customer.area,
            'address': customer.address,
            'tasks': _trimmed(request.data.get('tasks')),
            'definition': definition,
            'status': 'in_progress',
            'due_date': request.data.get('due_date') or None,
            'revenue': _trimmed(request.data.get('revenue')),
            'responsibility': _trimmed(request.data.get('responsibility')),
            'remark': _trimmed(request.data.get('remark')),
            'actions': [action_record],
        }

        serializer = WeeklyReportSerializer(data=report_data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        report = serializer.save()
        return Response(WeeklyReportSerializer(report).data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        queryset = self.filter_queryset(
            self.get_queryset().prefetch_related('contacts', 'revenue_records')
        )

        workbook = Workbook()
        _add_customers_sheet(workbook, queryset)
        return _xlsx_response(workbook, 'customers.xlsx')

    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        template_path = _customer_import_template_path()
        if not template_path.exists():
            return Response({'error': '客户导入模板不存在'}, status=status.HTTP_404_NOT_FOUND)

        response = HttpResponse(template_path.read_bytes(), content_type=EXCEL_CONTENT_TYPE)
        response['Content-Disposition'] = 'attachment; filename="customer_import_template.xlsx"'
        return response

    @action(
        detail=False,
        methods=['post'],
        url_path='import',
        parser_classes=[MultiPartParser, FormParser],
    )
    def import_file(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = _read_customer_import_rows(upload)
            result = _import_customer_and_contact_rows(rows)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(result)


class ContactViewSet(viewsets.ModelViewSet):
    """联系人视图集"""

    queryset = Contact.objects.all()
    serializer_class = ContactSerializer
    pagination_class = None
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['customer', 'is_key_person']
    search_fields = ['name', 'position', 'phone', 'email']

    def get_queryset(self):
        """优化查询性能"""
        return super().get_queryset().select_related('customer')


class CustomerRevenueViewSet(viewsets.ModelViewSet):
    """客户营收视图集"""

    queryset = CustomerRevenue.objects.all()
    serializer_class = CustomerRevenueSerializer
    pagination_class = None
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['customer', 'month']
    search_fields = ['customer__client_name']
    ordering_fields = ['month', 'revenue', 'created_at', 'updated_at']
    ordering = ['-month', 'customer__client_name']
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def get_queryset(self):
        return super().get_queryset().select_related('customer')

    @action(detail=False, methods=['post'], url_path='import')
    def import_file(self, request):
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            rows = _read_revenue_rows(upload)
        except ValueError as exc:
            return Response({'error': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        imported = 0
        updated = 0
        skipped = 0
        errors = []

        for row_number, row in rows:
            customer_name = (row.get('customer name') or '').strip()
            customer = Customer.objects.filter(client_name__iexact=customer_name).first()
            if not customer:
                skipped += 1
                errors.append({
                    'row': row_number,
                    'customer_name': customer_name,
                    'reason': '客户名称未完全匹配 Fishpool 客户'
                })
                continue

            try:
                month = _parse_month(row.get('month'))
                revenue = _parse_revenue(row.get('revenue'))
            except ValueError as exc:
                skipped += 1
                errors.append({
                    'row': row_number,
                    'customer_name': customer_name,
                    'reason': str(exc)
                })
                continue

            _, created = CustomerRevenue.objects.update_or_create(
                customer=customer,
                month=month,
                defaults={'revenue': revenue}
            )
            if created:
                imported += 1
            else:
                updated += 1

        return Response({
            'imported': imported,
            'updated': updated,
            'skipped': skipped,
            'errors': errors[:50],
        })


class WeeklyReportViewSet(viewsets.ModelViewSet):
    """Weekly Report 视图集"""

    queryset = WeeklyReport.objects.all()
    serializer_class = WeeklyReportSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['client_name', 'area', 'tasks', 'responsibility', 'customer', 'status']
    search_fields = ['client_name', 'definition', 'remark']
    ordering_fields = ['due_date', 'created_at', 'updated_at']
    ordering = ['-due_date', '-created_at']

    def get_serializer_class(self):
        """根据 action 选择序列化器"""
        if self.action == 'list':
            return WeeklyReportListSerializer
        return WeeklyReportSerializer

    def get_queryset(self):
        """优化查询性能，支持日期范围筛选"""
        queryset = super().get_queryset()
        if self.action in ['list', 'retrieve']:
            queryset = queryset.select_related('customer')

        # 日期范围筛选
        due_date_after = self.request.query_params.get('due_date_after')
        due_date_before = self.request.query_params.get('due_date_before')

        if due_date_after:
            queryset = queryset.filter(due_date__gte=due_date_after)
        if due_date_before:
            queryset = queryset.filter(due_date__lte=due_date_before)

        return queryset

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        queryset = self.filter_queryset(self.get_queryset().select_related('customer'))

        workbook = Workbook()
        _add_weekly_reports_sheet(workbook, queryset)
        return _xlsx_response(workbook, 'weekly-reports.xlsx')

    @action(detail=True, methods=['get', 'post'], url_path='actions')
    def actions_list(self, request, pk=None):
        """获取或创建行动记录"""
        report = self.get_object()

        if request.method == 'GET':
            # 获取行动记录列表
            actions_with_id = []
            if report.actions:
                for index, action in enumerate(report.actions):
                    action_data = {
                        'id': index,
                        'action_date': action.get('action_date', ''),
                        'action': action.get('action', action.get('content', '')),  # 兼容旧格式
                        'result': action.get('result', ''),
                        'next_step': action.get('next_step', ''),
                        'timestamp': action.get('timestamp', ''),
                        'user': action.get('user', '')
                    }
                    actions_with_id.append(action_data)
            return Response(actions_with_id)

        elif request.method == 'POST':
            # 创建行动记录
            action_date = request.data.get('action_date', '').strip()
            action_content = request.data.get('action', '').strip()
            result = request.data.get('result', '').strip()
            next_step = request.data.get('next_step', '').strip()

            if not action_content:
                return Response(
                    {'error': '行动内容不能为空'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 创建新的行动记录
            action_record = {
                'timestamp': timezone.now().isoformat(),
                'action_date': action_date,
                'action': action_content,
                'result': result,
                'next_step': next_step
            }

            # 添加到 actions 列表
            if report.actions is None:
                report.actions = []
            report.actions.append(action_record)
            report.save()

            # 返回新创建的记录（带 id）
            new_action = {
                'id': len(report.actions) - 1,
                **action_record
            }
            return Response(new_action, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['put', 'delete'], url_path='actions/(?P<action_id>[0-9]+)')
    def actions_detail(self, request, pk=None, action_id=None):
        """更新或删除行动记录（RESTful 风格）"""
        report = self.get_object()
        try:
            action_id = int(action_id)
        except (ValueError, TypeError):
            return Response(
                {'error': '无效的行动记录 ID'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not report.actions or action_id >= len(report.actions):
            return Response(
                {'error': '行动记录不存在'},
                status=status.HTTP_404_NOT_FOUND
            )

        if request.method == 'PUT':
            # 更新行动记录
            action_date = request.data.get('action_date', '').strip()
            action_content = request.data.get('action', '').strip()
            result = request.data.get('result', '').strip()
            next_step = request.data.get('next_step', '').strip()

            if not action_content:
                return Response(
                    {'error': '行动内容不能为空'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 更新行动记录（保留原始时间戳）
            report.actions[action_id].update({
                'action_date': action_date,
                'action': action_content,
                'result': result,
                'next_step': next_step
            })
            report.save()

            # 返回更新后的记录
            updated_action = {
                'id': action_id,
                **report.actions[action_id]
            }
            return Response(updated_action)

        elif request.method == 'DELETE':
            # 删除行动记录
            report.actions.pop(action_id)
            report.save()
            return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'], url_path='add-action')
    def add_action(self, request, pk=None):
        """添加行动记录（兼容旧格式）"""
        report = self.get_object()

        # 支持新旧两种格式
        action_date = request.data.get('action_date', '').strip()
        action_content = request.data.get('action', request.data.get('content', '')).strip()
        result = request.data.get('result', '').strip()
        next_step = request.data.get('next_step', '').strip()
        user = request.data.get('user', '').strip()

        if not action_content:
            return Response(
                {'error': '行动记录内容不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 创建新的行动记录
        action_record = {
            'timestamp': timezone.now().isoformat(),
            'action_date': action_date,
            'action': action_content,
            'result': result,
            'next_step': next_step,
            'user': user
        }

        # 添加到 actions 列表
        if report.actions is None:
            report.actions = []
        report.actions.append(action_record)
        report.save()

        serializer = self.get_serializer(report)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='actions/(?P<index>[0-9]+)/update')
    def update_action(self, request, pk=None, index=None):
        """编辑行动记录（兼容旧格式）"""
        report = self.get_object()
        try:
            index = int(index)
        except (ValueError, TypeError):
            return Response(
                {'error': '无效的索引'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not report.actions or index >= len(report.actions):
            return Response(
                {'error': '行动记录不存在'},
                status=status.HTTP_404_NOT_FOUND
            )

        # 支持新旧两种格式
        action_date = request.data.get('action_date', '').strip()
        action_content = request.data.get('action', request.data.get('content', '')).strip()
        result = request.data.get('result', '').strip()
        next_step = request.data.get('next_step', '').strip()
        user = request.data.get('user', '').strip()

        if not action_content:
            return Response(
                {'error': '行动记录内容不能为空'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 更新行动记录（保留原始时间戳）
        report.actions[index].update({
            'action_date': action_date,
            'action': action_content,
            'result': result,
            'next_step': next_step,
            'user': user
        })
        report.save()

        serializer = self.get_serializer(report)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='actions/(?P<index>[0-9]+)/delete')
    def delete_action(self, request, pk=None, index=None):
        """删除行动记录"""
        report = self.get_object()
        try:
            index = int(index)
        except (ValueError, TypeError):
            return Response(
                {'error': '无效的索引'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not report.actions or index >= len(report.actions):
            return Response(
                {'error': '行动记录不存在'},
                status=status.HTTP_404_NOT_FOUND
            )

        # 删除行动记录
        report.actions.pop(index)
        report.save()

        serializer = self.get_serializer(report)
        return Response(serializer.data)
