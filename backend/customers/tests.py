from rest_framework import status
from rest_framework.test import APITestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from io import BytesIO
import json
from openpyxl import Workbook, load_workbook

from .models import Contact, Customer, CustomerRevenue, WeeklyReport, get_last_quarter_range


class CustomerAndContactApiTests(APITestCase):
    def setUp(self):
        self.customer_a = Customer.objects.create(
            client_name='Area A Customer',
            business_model='Hunting',
            area='华东',
            city='上海',
        )
        self.customer_b = Customer.objects.create(
            client_name='Area B Customer',
            business_model='Farming',
            area='华北',
            city='北京',
        )

    def test_customer_list_filters_by_area(self):
        response = self.client.get('/api/customers/', {'area': '华北'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['name'], self.customer_b.client_name)

    def test_customer_list_applies_custom_text_choice_and_number_filters(self):
        self.customer_a.alias = 'Alpha'
        self.customer_a.score_x = 80
        self.customer_a.potential_contribution = '120.00'
        self.customer_a.save()
        self.customer_b.alias = 'Beta'
        self.customer_b.score_x = 40
        self.customer_b.potential_contribution = '20.00'
        self.customer_b.save()

        filters = [
            {'field': 'alias', 'operator': 'contains', 'value': 'alp'},
            {'field': 'business_model', 'operator': 'eq', 'value': 'Hunting'},
            {'field': 'score_x', 'operator': 'gte', 'value': '70'},
            {'field': 'potential_contribution', 'operator': 'between', 'value': ['100', '130']},
        ]
        response = self.client.get('/api/customers/', {'filters': json.dumps(filters)})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['name'], self.customer_a.client_name)

    def test_customer_list_applies_custom_computed_revenue_filters(self):
        today = timezone.localdate()
        last_year_month = today.replace(year=today.year - 1, month=5, day=1)
        last_quarter_start, _ = get_last_quarter_range(today)

        CustomerRevenue.objects.create(
            customer=self.customer_a,
            month=last_year_month,
            revenue='1200.00',
        )
        CustomerRevenue.objects.create(
            customer=self.customer_a,
            month=last_quarter_start,
            revenue='300.00',
        )
        CustomerRevenue.objects.create(
            customer=self.customer_b,
            month=last_year_month,
            revenue='500.00',
        )
        CustomerRevenue.objects.create(
            customer=self.customer_b,
            month=last_quarter_start,
            revenue='900.00',
        )

        filters = [
            {'field': 'last_year_revenue', 'operator': 'gte', 'value': '1000'},
            {'field': 'last_quarter_revenue', 'operator': 'lte', 'value': '500'},
        ]
        response = self.client.get('/api/customers/', {'filters': json.dumps(filters)})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 1)
        self.assertEqual(response.data['results'][0]['name'], self.customer_a.client_name)

    def test_customer_list_orders_by_custom_filter_field(self):
        self.customer_a.alias = 'Alpha'
        self.customer_a.save()
        self.customer_b.alias = 'Zulu'
        self.customer_b.save()

        response = self.client.get('/api/customers/', {'ordering': 'alias'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['results'][0]['name'], self.customer_a.client_name)
        self.assertEqual(response.data['results'][1]['name'], self.customer_b.client_name)

    def test_customer_list_orders_by_computed_revenue_field(self):
        last_quarter_start, _ = get_last_quarter_range(timezone.localdate())
        CustomerRevenue.objects.create(
            customer=self.customer_a,
            month=last_quarter_start,
            revenue='300.00',
        )
        CustomerRevenue.objects.create(
            customer=self.customer_b,
            month=last_quarter_start,
            revenue='900.00',
        )

        response = self.client.get('/api/customers/', {'ordering': 'last_quarter_revenue'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['results'][0]['name'], self.customer_a.client_name)
        self.assertEqual(response.data['results'][1]['name'], self.customer_b.client_name)

    def test_customer_list_respects_page_size_query_param(self):
        for index in range(23):
            Customer.objects.create(
                client_name=f'Paged Customer {index}',
                business_model='Hunting',
            )

        response = self.client.get('/api/customers/', {'page': 3, 'page_size': 12})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['count'], 25)
        self.assertEqual(len(response.data['results']), 1)

    def test_contact_list_is_not_paginated(self):
        for index in range(21):
            Contact.objects.create(
                customer=self.customer_a,
                name=f'Contact {index}',
            )

        response = self.client.get('/api/contacts/', {'customer': self.customer_a.id})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIsInstance(response.data, list)
        self.assertEqual(len(response.data), 21)

    def test_create_weekly_report_allows_blank_optional_dates(self):
        payload = {
            'client_name': self.customer_a.client_name,
            'area': self.customer_a.area,
            'address': 'Test address',
            'tasks': 'Task A',
            'definition': 'Definition A',
            'due_date': '2026-04-20',
            'revise_date': '',
            'finish_date': '',
            'revenue': '100',
            'responsibility': 'Tester',
            'remark': 'note',
            'actions': [],
        }

        response = self.client.post('/api/weekly-reports/', payload, format='json')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        report = WeeklyReport.objects.get(client_name=self.customer_a.client_name, definition='Definition A')
        self.assertIsNone(report.revise_date)
        self.assertIsNone(report.finish_date)

    def test_create_customer_revenue_matches_customer_name_case_insensitively(self):
        response = self.client.post('/api/customer-revenues/', {
            'customer_name': self.customer_a.client_name.upper(),
            'month': '2026-01-01',
            'revenue': '1000.00',
        }, format='json')

        self.assertEqual(response.status_code, status.HTTP_201_CREATED)
        self.assertEqual(CustomerRevenue.objects.count(), 1)
        self.assertEqual(CustomerRevenue.objects.get().customer, self.customer_a)

    def test_import_customer_revenue_skips_unmatched_customers(self):
        csv_content = (
            'month,customer name,revenue\n'
            f'2026-01,{self.customer_a.client_name},1000\n'
            '2026-01,Unknown Customer,2000\n'
        ).encode('utf-8')
        upload = SimpleUploadedFile('revenues.csv', csv_content, content_type='text/csv')

        response = self.client.post('/api/customer-revenues/import/', {
            'file': upload,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 1)
        self.assertEqual(response.data['skipped'], 1)
        self.assertEqual(CustomerRevenue.objects.count(), 1)
        self.assertEqual(CustomerRevenue.objects.get().customer, self.customer_a)

    def test_import_customer_revenue_reads_all_xlsx_sheets(self):
        workbook = Workbook()
        sheet_2021 = workbook.active
        sheet_2021.title = '2021'
        sheet_2021.append(['month', 'customer name', 'revenue'])
        sheet_2021.append(['2021-01', self.customer_a.client_name, 1000])

        sheet_2024 = workbook.create_sheet('2024')
        sheet_2024.append(['month', 'customer name', 'revenue'])
        sheet_2024.append(['2024-01', self.customer_a.client_name, 4000])
        workbook.active = 1

        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)
        upload = SimpleUploadedFile(
            'revenues.xlsx',
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post('/api/customer-revenues/import/', {
            'file': upload,
        }, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['imported'], 2)
        self.assertEqual(CustomerRevenue.objects.count(), 2)
        self.assertTrue(CustomerRevenue.objects.filter(month='2021-01-01', revenue='1000.00').exists())
        self.assertTrue(CustomerRevenue.objects.filter(month='2024-01-01', revenue='4000.00').exists())

    def test_customer_payload_includes_last_year_and_last_quarter_revenue(self):
        today = timezone.localdate()
        last_year_month = today.replace(year=today.year - 1, month=5, day=1)
        last_quarter_start, _ = get_last_quarter_range(today)

        CustomerRevenue.objects.create(
            customer=self.customer_a,
            month=last_year_month,
            revenue='1200.00',
        )
        CustomerRevenue.objects.create(
            customer=self.customer_a,
            month=last_quarter_start,
            revenue='300.00',
        )

        response = self.client.get(f'/api/customers/{self.customer_a.id}/')

        expected_last_year = '1500.00' if last_quarter_start.year == today.year - 1 else '1200.00'
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['last_year_revenue'], expected_last_year)
        self.assertEqual(response.data['last_quarter_revenue'], '300.00')

    def test_customer_revenue_summary_groups_monthly_revenue_by_year(self):
        CustomerRevenue.objects.create(customer=self.customer_a, month='2021-01-01', revenue='1200.00')
        CustomerRevenue.objects.create(customer=self.customer_a, month='2021-02-01', revenue='1500.00')
        CustomerRevenue.objects.create(customer=self.customer_a, month='2022-01-01', revenue='1800.00')
        CustomerRevenue.objects.create(customer=self.customer_b, month='2021-01-01', revenue='9999.00')

        response = self.client.get(f'/api/customers/{self.customer_a.id}/revenue-summary/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['yearly'], [
            {'year': 2021, 'revenue': '2700.00'},
            {'year': 2022, 'revenue': '1800.00'},
        ])
        self.assertEqual(response.data['monthly'], [
            {'month': '2021-01', 'revenue': '1200.00'},
            {'month': '2021-02', 'revenue': '1500.00'},
            {'month': '2022-01', 'revenue': '1800.00'},
        ])

    def test_customer_revenue_summary_filters_monthly_rows_by_year(self):
        CustomerRevenue.objects.create(customer=self.customer_a, month='2021-01-01', revenue='1200.00')
        CustomerRevenue.objects.create(customer=self.customer_a, month='2022-01-01', revenue='1800.00')

        response = self.client.get(f'/api/customers/{self.customer_a.id}/revenue-summary/', {'year': '2022'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['yearly'], [
            {'year': 2021, 'revenue': '1200.00'},
            {'year': 2022, 'revenue': '1800.00'},
        ])
        self.assertEqual(response.data['monthly'], [
            {'month': '2022-01', 'revenue': '1800.00'},
        ])

    def test_customer_revenue_summary_returns_empty_arrays_without_revenue(self):
        response = self.client.get(f'/api/customers/{self.customer_a.id}/revenue-summary/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['yearly'], [])
        self.assertEqual(response.data['monthly'], [])

    def test_export_customers_respects_filters(self):
        response = self.client.get('/api/customers/export/', {'area': self.customer_b.area})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(rows[0][0], '客户名称')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], self.customer_b.client_name)

    def test_export_weekly_reports_respects_filters(self):
        WeeklyReport.objects.create(
            client_name=self.customer_a.client_name,
            customer=self.customer_a,
            area=self.customer_a.area,
            tasks='Task A',
            definition='Definition A',
            responsibility='Tester A',
            status='in_progress',
        )
        WeeklyReport.objects.create(
            client_name=self.customer_b.client_name,
            customer=self.customer_b,
            area=self.customer_b.area,
            tasks='Task B',
            definition='Definition B',
            responsibility='Tester B',
            status='completed',
        )

        response = self.client.get('/api/weekly-reports/export/', {'status': 'completed'})

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(rows[0][0], '客户名称')
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][0], self.customer_b.client_name)
        self.assertEqual(rows[1][6], '已完成')

    def test_export_workbook_contains_customers_and_weekly_report_actions(self):
        WeeklyReport.objects.create(
            client_name=self.customer_a.client_name,
            customer=self.customer_a,
            area=self.customer_a.area,
            address='Test address',
            tasks='Task A',
            definition='Definition A',
            responsibility='Tester A',
            status='in_progress',
            actions=[
                {
                    'action_date': '2026-04-01',
                    'action': 'First action',
                    'result': 'First result',
                    'next_step': 'First next step',
                    'user': 'User A',
                    'timestamp': '2026-04-01T10:00:00',
                },
                {
                    'action_date': '2026-04-02',
                    'content': 'Second action',
                    'result': 'Second result',
                    'next_step': 'Second next step',
                    'user': 'User B',
                    'timestamp': '2026-04-02T10:00:00',
                },
            ],
        )

        response = self.client.get('/api/export/workbook/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ['客户池', '周报'])

        customer_rows = list(workbook['客户池'].iter_rows(values_only=True))
        weekly_rows = list(workbook['周报'].iter_rows(values_only=True))
        workbook.close()

        self.assertEqual(customer_rows[0][0], '客户名称')
        self.assertEqual(len(customer_rows), 3)
        self.assertEqual(weekly_rows[0][15], 'Action内容')
        self.assertEqual(len(weekly_rows), 3)
        self.assertEqual(weekly_rows[1][15], 'First action')
        self.assertEqual(weekly_rows[2][15], 'Second action')
