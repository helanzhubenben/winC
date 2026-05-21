from io import BytesIO
from pathlib import Path

from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework import status
from rest_framework.test import APITestCase
from openpyxl import Workbook, load_workbook

from customers.models import Contact, Customer


HEADERS = [
    '客户名称',
    '业务模式',
    '区域',
    '城市',
    '地址',
    'X轴描述',
    'X轴评分',
    'Y轴描述',
    'Y轴评分',
    'Z轴描述',
    'Z轴评分',
    '客户策略',
    '潜在贡献',
    '备注',
    '联系人姓名',
    '联系人职位',
    '联系人电话',
    '联系人邮箱',
    '是否关键人',
]


def make_import_upload(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '客户导入'
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return SimpleUploadedFile(
        'customers.xlsx',
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


class CustomerImportApiTests(APITestCase):
    def test_download_customer_import_template_matches_single_sheet_spec(self):
        response = self.client.get('/api/customers/import-template/')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        self.assertEqual(workbook.sheetnames, ['客户导入'])
        headers = [cell.value for cell in next(workbook['客户导入'].iter_rows(min_row=1, max_row=1))]
        workbook.close()
        self.assertEqual(headers, HEADERS)

    def test_import_creates_customer_contacts_and_keeps_first_duplicate_customer_values(self):
        upload = make_import_upload([
            [
                '导入客户A',
                'Hunting',
                '华东',
                '上海',
                '上海市测试路1号',
                '客户潜力高',
                85,
                '竞争环境可控',
                75,
                '关键人关系较好',
                75,
                '重点跟进',
                180.50,
                '第一行客户字段应保留',
                '张三',
                '采购经理',
                '13800000001',
                'zhangsan@example.com',
                '是',
            ],
            [
                '导入客户A',
                'Farming',
                '华北',
                '北京',
                '不应覆盖的地址',
                '不应覆盖',
                10,
                '不应覆盖',
                20,
                '不应覆盖',
                30,
                '不应覆盖',
                1,
                '第二行只应处理联系人',
                '李四',
                '质量总监',
                '13800000002',
                'lisi@example.com',
                '否',
            ],
            [
                '导入客户B',
                'Farming',
                '西南',
                '成都',
                '成都市测试路2号',
                '存量客户',
                55,
                '竞争强',
                45,
                '关系一般',
                35,
                '维护',
                '',
                '无联系人行',
                '',
                '',
                '',
                '',
                '',
            ],
        ])

        response = self.client.post('/api/customers/import/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['customers_created'], 2)
        self.assertEqual(response.data['customers_updated'], 0)
        self.assertEqual(response.data['contacts_created'], 2)
        self.assertEqual(response.data['contacts_updated'], 0)
        self.assertEqual(response.data['skipped'], 0)
        self.assertEqual(response.data['errors'], [])
        self.assertEqual(len(response.data['warnings']), 1)

        customer_a = Customer.objects.get(client_name='导入客户A')
        self.assertEqual(customer_a.business_model, 'Hunting')
        self.assertEqual(customer_a.area, '华东')
        self.assertEqual(customer_a.city, '上海')
        self.assertEqual(customer_a.address, '上海市测试路1号')
        self.assertEqual(customer_a.score_x, 85)
        self.assertEqual(customer_a.score_y, 75)
        self.assertEqual(customer_a.score_z, 75)
        self.assertEqual(customer_a.level, 'A')
        self.assertEqual(str(customer_a.potential_contribution), '180.50')
        self.assertEqual(customer_a.contacts.count(), 2)
        self.assertTrue(Contact.objects.get(customer=customer_a, name='张三').is_key_person)
        self.assertFalse(Contact.objects.get(customer=customer_a, name='李四').is_key_person)

        customer_b = Customer.objects.get(client_name='导入客户B')
        self.assertEqual(customer_b.contacts.count(), 0)
        self.assertIsNone(customer_b.potential_contribution)

    def test_import_updates_existing_customer_and_contact_by_exact_names(self):
        customer = Customer.objects.create(
            client_name='已有客户',
            business_model='Hunting',
            area='旧区域',
            city='旧城市',
            score_x=10,
            score_y=20,
            score_z=30,
            level='D',
        )
        Contact.objects.create(
            customer=customer,
            name='王五',
            position='旧职位',
            phone='10000000000',
            email='old@example.com',
            is_key_person=False,
        )
        upload = make_import_upload([
            [
                '已有客户',
                'Farming',
                '华南',
                '广州',
                '广州市测试路3号',
                '更新后的X描述',
                80,
                '更新后的Y描述',
                80,
                '更新后的Z描述',
                40,
                '更新后的策略',
                88.80,
                '更新后的备注',
                '王五',
                '新职位',
                '13800000005',
                'new@example.com',
                'Y',
            ],
        ])

        response = self.client.post('/api/customers/import/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['customers_created'], 0)
        self.assertEqual(response.data['customers_updated'], 1)
        self.assertEqual(response.data['contacts_created'], 0)
        self.assertEqual(response.data['contacts_updated'], 1)

        customer.refresh_from_db()
        self.assertEqual(customer.business_model, 'Farming')
        self.assertEqual(customer.area, '华南')
        self.assertEqual(customer.city, '广州')
        self.assertEqual(customer.level, 'B')
        self.assertEqual(str(customer.potential_contribution), '88.80')

        contact = Contact.objects.get(customer=customer, name='王五')
        self.assertEqual(contact.position, '新职位')
        self.assertEqual(contact.phone, '13800000005')
        self.assertEqual(contact.email, 'new@example.com')
        self.assertTrue(contact.is_key_person)

    def test_import_accepts_sample_customer_import_test_data_file(self):
        sample_path = Path(__file__).resolve().parents[2] / 'docs' / 'templates' / 'customer_import_test_data.xlsx'
        upload = SimpleUploadedFile(
            'customer_import_test_data.xlsx',
            sample_path.read_bytes(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post('/api/customers/import/', {'file': upload}, format='multipart')

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['customers_created'], 4)
        self.assertEqual(response.data['contacts_created'], 4)
        self.assertEqual(response.data['skipped'], 0)
        self.assertEqual(Customer.objects.count(), 4)
        self.assertEqual(Contact.objects.count(), 4)
