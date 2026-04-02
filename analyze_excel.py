#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""分析销售周报Excel文件"""

from openpyxl import load_workbook
import json

def analyze_excel(file_path):
    """分析Excel文件结构"""
    wb = load_workbook(file_path, data_only=True)

    result = {
        "工作表数量": len(wb.sheetnames),
        "工作表列表": wb.sheetnames,
        "详细分析": {}
    }

    # 分析关键工作表
    key_sheets = ['General-Fishpool', 'Weekly Report', 'M & S Key-Customers', ' Top Clients']

    for sheet_name in key_sheets:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_info = {
                "行数": sheet.max_row,
                "列数": sheet.max_column,
                "表头": [],
                "样本数据": []
            }

            # 读取前2行作为表头
            for row_idx in range(1, min(3, sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(sheet.max_column + 1, 20)):
                    cell_value = sheet.cell(row_idx, col_idx).value
                    if cell_value:
                        row_data.append(str(cell_value))
                sheet_info["表头"].append(row_data)

            # 读取3-5行作为样本
            for row_idx in range(3, min(6, sheet.max_row + 1)):
                row_data = []
                for col_idx in range(1, min(sheet.max_column + 1, 15)):
                    cell_value = sheet.cell(row_idx, col_idx).value
                    if cell_value:
                        val_str = str(cell_value)[:50]
                        row_data.append(val_str)
                if row_data:
                    sheet_info["样本数据"].append(row_data)

            result["详细分析"][sheet_name] = sheet_info

    return result

if __name__ == "__main__":
    file_path = r"D:\BaiduSyncdisk\TRIGO\TRIGO_CHINA\02_GRB\00-Sales\Weekly-report\★ 2026Fish Pool & Weekly report-Timothee.xlsx"

    result = analyze_excel(file_path)

    print("=" * 80)
    print("Excel文件分析报告")
    print("=" * 80)
    print(f"\n工作表数量: {result['工作表数量']}")
    print(f"\n工作表列表:")
    for i, name in enumerate(result['工作表列表'], 1):
        print(f"  {i}. {name}")

    print("\n" + "=" * 80)
    print("关键工作表详细分析")
    print("=" * 80)

    for sheet_name, info in result['详细分析'].items():
        print(f"\n【{sheet_name}】")
        print(f"  规模: {info['行数']}行 x {info['列数']}列")

        print(f"\n  表头结构:")
        for i, header_row in enumerate(info['表头'], 1):
            if header_row:
                print(f"    第{i}行: {' | '.join(header_row[:10])}")

        print(f"\n  数据样本:")
        for i, data_row in enumerate(info['样本数据'], 1):
            if data_row:
                print(f"    样本{i}: {' | '.join(data_row[:5])}")

        print("-" * 80)
